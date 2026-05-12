from fastapi import APIRouter, Query, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import io
from rich import print
from reporting.db.mongo import get_reporting_collection

router = APIRouter()

async def secondes_to_timedelta(total_seconds):
    return timedelta(seconds=total_seconds)

@router.get("/excel")
async def generate_excel_report(
    date_debut: str = Query(..., description="Date de début (format: YYYY-MM-DD)"),
    date_fin: str = Query(..., description="Date de fin (format: YYYY-MM-DD)")
):
    try:
        try:
            start_dt = datetime.strptime(date_debut, "%Y-%m-%d")
            end_dt = datetime.strptime(date_fin, "%Y-%m-%d") + timedelta(days=1)
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail="Format de date invalide. Utilisez YYYY-MM-DD."
            )
        collection = get_reporting_collection()
        query = {"Database_date": {"$gte": start_dt, "$lt": end_dt}}
        cursor = collection.find(query, {"_id": 0})
        docs = await cursor.to_list(length=10000)
        docs.sort(key=lambda x: (str(x.get("Vehicules", "")).strip().upper()))

        if not docs:
            raise HTTPException(
                status_code=404,
                detail="Aucune donnée trouvée pour la période spécifiée."
            )

        # En-têtes de base
        headers_base = [
            "Date", "Marque", "Voiture", "Immatriculation", "Kilométrage du véhicule (km)",
            "Km parcouru", "Quantité", "Temp de Service", "Temps d’arrêt pendant les livraisons"
        ]

        # Styles communs
        title_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        cell_font = Font(name="Calibri", size=10)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))
        centered = Alignment(horizontal="center", vertical="center")

        # Préparation des données
        thermique_rows = []
        electrique_rows = []

        for doc in docs:
            service_time_sec = doc.get('Service', 0)
            stop_service_sec = doc.get('Stop_service', 0)

            if service_time_sec == 0 and stop_service_sec == 0:
                continue

            service_time = await secondes_to_timedelta(service_time_sec)
            stop_service = await secondes_to_timedelta(stop_service_sec)
            fuel_liters = int(doc.get('fuel', 0)) / 1000

            base_row = [
                doc.get("Date", ""),
                doc.get("Marque", ""),
                str(doc.get("Vehicules", "")).upper(),
                doc.get("Immatriculation", ""),
                doc.get("odometer", ""),
                doc.get('distance'),
                None,
                service_time,
                stop_service,
            ]

            if fuel_liters == 0.0:
                percentage = 0
                row = base_row.copy()
                row[6] = f"{percentage:.2f}%"
                electrique_rows.append(row)
            else:
                row = base_row.copy()
                row[6] = f"{fuel_liters}L"
                thermique_rows.append(row)
        
        wb = Workbook()
        if wb.sheetnames:
            wb.remove(wb.active)

        # Date de génération
        generate_date = ""
        for doc in reversed(docs):
            db_date = doc.get("Database_date")
            if db_date:
                generate_date = db_date.strftime("%d/%m/%Y %H:%M:%S")
                break

        def apply_styles(ws, header_row=7, num_cols=9):
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=header_row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = centered
                cell.border = thin_border
            for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, min_col=1, max_col=num_cols):
                for cell in row:
                    cell.font = cell_font
                    cell.border = thin_border
                    cell.alignment = centered
                    if cell.column == 5 and isinstance(cell.value, (int, float)):
                        cell.number_format = "#,##0.00"

        def setup_sheet(ws, title_text, rows, is_electric):
            # Titre principal
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers_base))
            if start_dt == end_dt - timedelta(days=1):
                ws["A1"].value = title_text
                ws.cell(row=3, column=1, value="Date et heure : ")
                ws.cell(row=3, column=2, value=generate_date)
            else:
                ws["A1"].value = f"{title_text} du {start_dt.strftime('%d/%m/%Y')} au {(end_dt - timedelta(days=1)).strftime('%d/%m/%Y')}"
                ws.cell(row=3, column=1, value="Date de Début : ")
                ws.cell(row=3, column=2, value=f"{start_dt.strftime('%d/%m/%Y %H:%M:%S')}")
                ws.cell(row=4, column=1, value="Date de Fin : ")
                ws.cell(row=4, column=2, value=generate_date)

            ws["A1"].font = Font(name="Calibri", bold=True, size=14)
            ws["A1"].fill = title_fill
            ws["A1"].alignment = centered

            # En-têtes
            headers = headers_base.copy()
            if is_electric:
                headers[6] = "Pourcentage de Batterie"
            
            for col_num, header in enumerate(headers, start=1):
                ws.cell(row=7, column=col_num, value=header)

            # Écriture des lignes
            for row_data in rows:
                ws.append(row_data)
            
            # Application des styles
            apply_styles(ws, num_cols=len(headers))

        if thermique_rows:
            ws_thermique = wb.create_sheet(title="Thermique")
            setup_sheet(ws_thermique, "Historique dotation carburant", thermique_rows, is_electric=False)
        
        if electrique_rows:
            ws_electrique = wb.create_sheet(title="Electrique")
            setup_sheet(ws_electrique, "Historique activité véhicules électriques", electrique_rows, is_electric=True)

        # Sauvegarde dans le buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        filename = f"Rapport_livraison_{date_debut}_au_{date_fin}.xlsx"
        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur serveur: {str(e)}")